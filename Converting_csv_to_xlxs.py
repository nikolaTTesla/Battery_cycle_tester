import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

def convert_csv_to_xlxs(exact_file_name, where_to_save):

    # Read CSV file
    #data = pd.read_csv('random_data.csv')
    data = pd.read_csv(where_to_save+'/'+exact_file_name)
    ##@@@@@@@@@@@@@@@@@@# Read the CSV file into a DataFrame
    # Write the DataFrame to an Excel file
    newFileLocation = where_to_save + '/'+ exact_file_name[:-4]+'_GRAPH_INCLUDED'+'.xlsx'
    data.to_excel(newFileLocation, index=True)
    data = pd.read_excel(newFileLocation)
    ##@@@@@@@@@@@@@@@@@@
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write data to Excel file
    for row_index, row in enumerate(data.values, start=1):
        for col_index, value in enumerate(row, start=1):
            if isinstance(value, str):
                try:
                    numeric_value = float(value)  # You can use int() for integers
                except ValueError:
                    numeric_value = value  # If it can't be converted, keep it as text
                ws.cell(row=row_index, column=col_index).value = numeric_value
            else:
                ws.cell(row=row_index, column=col_index).value = value



    # Create scatter chart
    chart = ScatterChart()
    chart.title = "Battery Test Discharge"
    chart.x_axis.title = 'Cycle'
    chart.y_axis.title = 'Capacity'
    # Set the size of the chart
    chart.width = 50  # Width in characters (approximate)
    chart.height = 20  # Height in rows (approximate)

    '''
    y_values_1 = Reference(ws, min_col=1, min_row=2, max_row=len(data), max_col=1)
    x_values_1 = Reference(ws, min_col=2, min_row=3, max_row=len(data), max_col=2)
    series_1 = Series(y_values_1, x_values_1, title_from_data=True)
    chart.series.append(series_1)
    '''
    counter = 1
    while counter < 9:
        min_max_col_var_y = 1 + (counter - 1) * 6
        min_max_col_var_x = 2 + (counter - 1) * 6
        y_values = Reference(ws, min_col=min_max_col_var_y, min_row=2, max_row=len(data), max_col=min_max_col_var_y)
        x_values = Reference(ws, min_col=min_max_col_var_x, min_row=3, max_row=len(data), max_col=min_max_col_var_x)
        series = Series(y_values, x_values, title_from_data=True)
        chart.series.append(series)
        counter = counter + 1


    # Add chart to worksheet
    location = "C"+str(len(data) +3)
    #ws.add_chart(chart, "C6")
    ws.add_chart(chart, location)


    # Save the workbook
    #wb.save('random_data_with_chart.xlsx')
    #wb.save(where_to_save + '/random_data_with_chart.xlsx')
    wb.save(newFileLocation)
